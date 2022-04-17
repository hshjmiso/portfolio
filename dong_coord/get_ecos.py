import requests
import time
from openpyxl import Workbook, load_workbook

FILE_PATH = 'C:/Users/My/Desktop/bdc'
STATISTIC_TABLE = '서비스 통계 목록'
STATISTIC_ITEM = '통계 세부항목 목록'
STATISTIC_SEARCH = '통계 조회'
STATISTIC_SEARCH_YY = '통계 조회 주기(년)'
STATISTIC_SEARCH_QQ = '통계 조회 주기(분기)'
STATISTIC_SEARCH_MM = '통계 조회 주기(월)'

def getURL(serviceName, startIndex, endIndex):
    URL = 'http://ecos.bok.or.kr/api'
    key = '5OSCO8U9FP9JUBPOCB96'
    return f'{URL}/{serviceName}/{key}/json/kr/{startIndex}/{endIndex}'

def callStatisticTableList(startIndex, endIndex):
    # 서비스 통계 목록
    # P_STAT_CODE	8	000Y074	                        상위통계표코드
    # STAT_CODE	    8	000Y702	                        통계표코드
    # STAT_NAME	    200	1.2.2 본원통화 구성내역(평잔)	  통계명
    # CYCLE	        2	YY, QQ, MM, DD	                주기(년, 분기, 월, 일)
    # SRCH_         YN	1	                            Y/N	검색가능여부
    # ORG_NAME	    50	국제통화기금(IMF)	             출처
    # http://ecos.bok.or.kr/api/StatisticTableList/5OSCO8U9FP9JUBPOCB96/json/kr/1/1/?/

    url = getURL('StatisticTableList', startIndex, endIndex) + '/?/'
    return requests.get(url).json()['StatisticTableList']


def callStatisticItemList(startIndex, endIndex, statCode):
    # 통계 세부항목 목록
    # STAT_CODE	    8	043Y070	                                통계표코드
    # STAT_NAME	    200	5.8.2 지역별 소비유형별 개인 신용카드	   통계명
    # GRP_CODE	    20	Group1	                                통계항목그룹코드
    # GRP_NAME	    60	지역코드	                             통계항목그룹명
    # P_ITEM_CODE	8	043Y070	                                상위통계항목코드
    # ITEM_CODE	    20	X	                                    통계항목코드
    # ITEM_NAME	    200	전국	                                통계항목명
    # CYCLE	        2	YY, QQ, MM, DD	                        주기(년, 분기, 월, 일)
    # START_TIME	8	200912	                                수록시작일자
    # END_TIME	    8	202005	                                수록종료일자
    # DATA_CNT	    22	10332	                                자료수
    # WEIGHT	    22	999.99	                                가중치
    # http://ecos.bok.or.kr/api/StatisticItemList/sample/xml/kr/1/10/?/
    
    url = f'{getURL("StatisticItemList", startIndex, endIndex)}/{statCode}/'
    return requests.get(url).json()['StatisticItemList']

def callStatisticSearch(startIndex, endIndex, statCode, cycle, startTime, endTime, itemCode):
    # 통계 조회 조건 설정
    # STAT_CODE	    8	027Y215	                    통계표코드
    # STAT_NAME	    200	12.15.1 성장성에 관한 지표	  통계명
    # ITEM_CODE1	20	1000	                    통계항목1코드
    # ITEM_NAME1	200	전산업	                    통계항목명1
    # ITEM_CODE2	20	A	                        통계항목2코드
    # ITEM_NAME2	200	종합	                    통계항목명2
    # ITEM_CODE3	20	1000	                    통계항목3코드
    # ITEM_NAME3	200	총자산증가율                통계항목명3
    # UNIT_NAME	    200	십억원	                    단위
    # TIME	        8	201101	                    시점
    # DATA_VALUE	23	48083.1	                    값
    # http://ecos.bok.or.kr/api/StatisticSearch/sample/xml/kr/1/10/010Y002/MM/201101/201101/AAAA11/

    url = f'{getURL("StatisticSearch", startIndex, endIndex)}/{statCode}/{cycle}/{startTime}/{endTime}/{itemCode}/'
    json = requests.get(url).json()

    if 'StatisticSearch' in json:
        print('hasattr')
        return json['StatisticSearch']
    else:
        print("hasn't attr")
        return {'row': [], 'list_total_count': 0}

# Write Excel 서비스 통계 목록 
def writeExcelStatisticTable():
    statisticTable = callStatisticTableList(1, 1)
    list_total_count = statisticTable['list_total_count']
    statisticTableList = callStatisticTableList(1, list_total_count)['row']

    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws.append(['통계명', '통계표코드'])

    for index, table in enumerate(statisticTableList):
        srch_yn = table['SRCH_YN']

        if srch_yn == 'Y':
            stat_name = table['STAT_NAME']
            stat_code = table['STAT_CODE']
            write_ws.append([stat_name, stat_code])
        
        print(f'writeExcelStatisticTable{index}')

    write_wb.save(f'{FILE_PATH}/서비스 통계 목록.xlsx')

def writeExcel(filename):
    write_wb = Workbook()
    write_wb.save(f'{FILE_PATH}/{filename}.xlsx')

def loadExcel(filename, sheet = 'Sheet'):
    load_wd = load_workbook(f'{FILE_PATH}/{filename}.xlsx', data_only=True)
    load_ws = load_wd[sheet]

    values = []
    for row in load_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        values.append(row_value)

    return [load_wd, load_ws, values]

# Write Excel 통계 세부항목 목록
def writeExcelStatisticItem(startIndex, endIndex):    
    [statisticTableWB, statisticTableWS, statisticTables] = loadExcel(STATISTIC_TABLE)

    [statisticItemWB, statisticItemWS, statisticItems] = loadExcel(STATISTIC_ITEM)

    if startIndex == 1:
        statisticItemWS.append(['통계명', '통계항목명', '통계표코드', '통계항목코드', '주기(년, 분기, 월, 일)', '수록시작일자', '수록종료일자'])
    for index, table in enumerate(statisticTables[ startIndex : endIndex ]):
        stat_name = table[0]
        stat_code = table[1]
        statisticItem = callStatisticItemList(1, 1, stat_code)
        list_total_count = statisticItem['list_total_count']
        statisticItemList = callStatisticItemList(1, list_total_count, stat_code)['row']
        for item in statisticItemList:
            stat_name = item['STAT_NAME']
            item_name = item['ITEM_NAME']
            stat_code = item['STAT_CODE']
            item_code = item['ITEM_CODE']
            cycle = item['CYCLE']
            start_time = item['START_TIME']
            end_time = item['END_TIME']
            
            statisticItemWS.append([stat_name, item_name, stat_code, item_code, cycle, start_time, end_time])
        
        print(startIndex + index)

    statisticItemWB.save(f'{FILE_PATH}/{STATISTIC_ITEM}.xlsx')


# Write Excel 통계 세부항목 목록
def writeExcelstatisticSearch(filename, dataList, cycle):
    [statisticItemWB, statisticItemWS, statisticItems] = loadExcel(STATISTIC_ITEM)

    [statisticSearchWB, statisticSearchWS, statisticSearches] = loadExcel(filename)
    for item in statisticItems:
        for data in dataList:
            if item[0].replace(" ", "") == data[0].replace(" ", "") and item[1].replace(" ", "") == data[1].replace(" ", ""):
                print(item)
                title = item[0].replace('/', '')
                subTitle = item[1].replace('/', '')
                sheet = f'{title}-{subTitle}'[0:31]
                write_ws = statisticSearchWB.create_sheet(sheet)
                stat_code = item[2]
                item_code = item[3]
                # cycle = item[4]
                start_time = item[5]
                end_time = item[6]
                statisticSearch = callStatisticSearch(1, 1, stat_code, cycle, start_time, end_time, item_code)
                write_ws.append(['통계표코드', '통계명', '통계항목1코드', '통계항목명1', '통계항목2코드', '통계항목명2', '통계항목3코드', '통계항목명3', '단위', '시점', '값'])
                # print(statisticSearch)
                list_total_count = statisticSearch['list_total_count']
                statisticSearchList = callStatisticSearch(1, list_total_count, stat_code, cycle, start_time, end_time, item_code)
                # print(statisticSearchList)
                for search in statisticSearchList['row']:
                    # print(search)
                    stat_code = search['STAT_CODE']
                    stat_name = search['STAT_NAME']
                    item_code1 = search['ITEM_CODE1']
                    item_name1 = search['ITEM_NAME1']
                    item_code2 = search['ITEM_CODE2']
                    item_name2 = search['ITEM_NAME2']
                    item_code3 = search['ITEM_CODE3']
                    item_name3 = search['ITEM_NAME3']
                    unit_name = search['UNIT_NAME']
                    times = search['TIME']
                    data_value = search['DATA_VALUE']
                    write_ws.append([stat_code, stat_name, item_code1, item_name1, item_code2, item_name2, item_code3, item_name3, unit_name, times, data_value])

                time.sleep(1)
            
    statisticSearchWB.save(f'{FILE_PATH}/{filename}.xlsx')





def run():
    print('시작')
    # writeExcelStatisticTable()
    # writeExcel('통계 세부항목 목록')
    # 596 ERROR!!!!
    # writeExcelStatisticItem(700, 701)

    # 주기 년도
    statListYY = [
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '4.1.2시장금리(월,분기,년)',
        '4.1.2시장금리(월,분기,년)',
        '17.거시경제분석지표',	
        '8.8.2.1주요국통화의대원화환율통계자료',	
        '6.1.2주식거래및주가지수',	
        '6.1.2주식거래및주가지수',
    ]

    itemListYY = [
        '경제성장률',
        '민간소비증감률(실질)',
        '설비투자증감률(실질)',
        '건설투자증감률(실질)',
        '국내총투자율',
        '제조업평균가동률',
        '생산자물가등락률',
        '소비자물가등락률',
        '수출물가등락률',
        '실업률',
        '고용률',
        'CD(91일)수익률',
        '국고채(3년)수익률',
        '회사채(장외3년,AA-등급)',
        '회사채(장외3년,BBB-등급)',
        'WTI현물유가등락률',
        '원/미국달러(매매기준율)',
        'KOSPI_평균',
        'KOSDAQ_평균'
    ]

    statListQQ = [
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	  
        '17.거시경제분석 지표',	
        '17.거시경제분석 지표',	
        '7.1.1 생산자물가지수(기본분류)(2015=100)',	
        '7.4.1 소비자물가지수(2015=100)(전국)',	
        '17.거시경제분석 지표',	  
        '17.거시경제분석지표',	
        '17.거시경제분석지표',	
        '4.1.2 시장금리(월,분기,년)',
        '4.1.2 시장금리(월,분기,년)',	
        '4.1.2 시장금리(월,분기,년)',	
        '4.1.2 시장금리(월,분기,년)',	
        '18.1.6.3 국제상품가격',	
        '8.8.2.1 주요국통화의 대원화 환율 통계자료',	
        '6.1.2 주식거래 및 주가지수',	  
        '3.7.3 은행대출금 연체율',	
        '3.7.3 은행대출금 연체율',	
        '17.거시경제분석 지표',
    ]
    
    itemListQQ = [
        '경제성장률',
        '민간소비증감률(실질)',
        '설비투자증감률(실질)',
        '건설투자증감률(실질)',
        '국내총투자율',
        '제조업 평균가동률',
        '경상수지',
        '상품수지',
        '서비스수지',
        '총지수',
        '총지수',
        '실업률',
        '주택매매가격등락률(전도시)',
        '주택전세가격등락률(전도시)',
        'CD유통수익률(91일)',
        '국고채(3년)',
        '회사채(장외3년,AA- 등급)',
        '회사채(장외3년,BBB-등급)',
        'WTI(현물)',
        '원/미국달러(매매기준율)',
        'KOSPI_평균',
        '기업대출-은행전체',
        '가계대출-은행전체',
        '가계신용잔액',
    ]	  

    statListMM = [
        '6.1.2 주식거래 및 주가지수',
        '17.거시경제분석 지표',
        '17.거시경제분석 지표',
    ]

    itemListMM = [
        'KOSPI_평균',
        '주택매매가격등락률(전도시)',
        '주택전세가격등락률(전도시)',
    ]

    # print(len(statListQQ))
    # print(len(itemListQQ))

    dataList = []

    for i, stat in enumerate(statListMM):
        dataList.append([statListMM[i], itemListMM[i]])

    print(dataList)
    
    writeExcel(STATISTIC_SEARCH_MM)
    writeExcelstatisticSearch(STATISTIC_SEARCH_MM, dataList, 'MM')
    


    # Excel 파일
    """
    [wd, ws, values] = loadExcel('통계 세부항목 목록_20211227_V0.1', '1차 샘플링 데이터')
    dataList = []

    for index, cell in enumerate(values[2]):
        dataList.append([values[2][index], values[3][index]])

    for index, data in enumerate(dataList):
        if not data[0] or not data[1]:
            del dataList[index]

    print(dataList)
    
    writeExcel(STATISTIC_SEARCH)
    writeExcelstatisticSearch(dataList)
    """
    
    """
    statisticTable = callStatisticTableList(1, 1)
    list_total_count = statisticTable['list_total_count']
    statisticTableList = callStatisticTableList(1, list_total_count)['row']
    # print(statisticTableList)
    
    # 전체 리스트 불러오기
    for table in statisticTableList:
        srch_yn = table['SRCH_YN']

        if srch_yn == 'Y':
            stat_code = table['STAT_CODE']
            statisticItem = callStatisticItemList(1, 1, stat_code)
            list_total_count = statisticItem['list_total_count']
            statisticItemList = callStatisticItemList(1, list_total_count, stat_code)['row']
            for item in statisticItemList:
                # callStatisticSearch(startIndex, endIndex, statCode, cycle, startTime, endTime)
                cycle = item['CYCLE']
                start_time = item['START_TIME']
                end_time = item['END_TIME']
                item_code = item['ITEM_CODE']
                statisticSearch = callStatisticSearch(1, 1, stat_code, cycle, start_time, end_time, item_code)
                list_total_count = statisticSearch['list_total_count']
                statisticSearchList = callStatisticSearch(1, list_total_count, stat_code, cycle, start_time, end_time, item_code)
                # print(statisticSearchList)
                time.sleep(1)
    """

    """
    # excel 파일 로드 (임시)
    load_wb = load_workbook("C:/Users/My/Desktop/통계리스트.xlsx", data_only=True)

    load_ws = load_wb['Sheet']

    all_values = []
    for row in load_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    print(all_values)
    """

    """
    # statisticTableListURL = '/StatisticTableList'
    # statisticItemListURL = '/StatisticItemList'
    # statisticSearchURL = '/StatisticSearch'
    

    print('시작')
    # 서비스 통계 목록
    # http://ecos.bok.or.kr/api/StatisticTableList/5OSCO8U9FP9JUBPOCB96/json/kr/1/1/?/
    # P_STAT_CODE	8	000Y074	상위통계표코드
    # STAT_CODE	8	000Y702	통계표코드
    # STAT_NAME	200	1.2.2 본원통화 구성내역(평잔)	통계명
    # CYCLE	2	YY, QQ, MM, DD	주기(년, 분기, 월, 일)
    # SRCH_YN	1	Y/N	검색가능여부
    # ORG_NAME	50	국제통화기금(IMF)	출처

    statisticTableListURL = f'{URL}/StatisticTableList/{key}/json/kr/1/1/?/'
    # print(statisticTableListURL)
    statisticTableList = requests.get(statisticTableListURL).json()['StatisticTableList']
    stat_code = statisticTableList['row'][1]['STAT_CODE']
    
    print(statisticTableList)
    print('\n')

    # 통계 세부항목 목록
    # http://ecos.bok.or.kr/api/StatisticItemList/sample/xml/kr/1/10/?/
    # STAT_CODE	8	043Y070	통계표코드
    # STAT_NAME	200	5.8.2 지역별 소비유형별 개인 신용카드	통계명
    # GRP_CODE	20	Group1	통계항목그룹코드
    # GRP_NAME	60	지역코드	통계항목그룹명
    # P_ITEM_CODE	8	043Y070	상위통계항목코드
    # ITEM_CODE	20	X	통계항목코드
    # ITEM_NAME	200	전국	통계항목명
    # CYCLE	2	YY, QQ, MM, DD	주기(년, 분기, 월, 일)
    # START_TIME	8	200912	수록시작일자
    # END_TIME	8	202005	수록종료일자
    # DATA_CNT	22	10332	자료수
    # WEIGHT	22	999.99	가중치

    statisticItemListURL = f'{URL}/StatisticItemList/{key}/json/kr/1/1/{stat_code}/'
    # print(statisticItemListURL)
    statisticItemList = requests.get(statisticItemListURL).json()['StatisticItemList']
    statisticItem = statisticItemList['row'][0]
    item_code = statisticItem['ITEM_CODE']
    start_time = statisticItem['START_TIME']
    end_time = statisticItem['END_TIME']
    cycle = statisticItem['CYCLE']

    print(statisticItemList)
    print('\n')

    # 통계 조회 조건 설정
    # http://ecos.bok.or.kr/api/StatisticSearch/sample/xml/kr/1/10/010Y002/MM/201101/201101/AAAA11/
    # STAT_CODE	8	027Y215	통계표코드
    # STAT_NAME	200	12.15.1 성장성에 관한 지표	통계명
    # ITEM_CODE1	20	1000	통계항목1코드
    # ITEM_NAME1	200	전산업	통계항목명1
    # ITEM_CODE2	20	A	통계항목2코드
    # ITEM_NAME2	200	종합	통계항목명2
    # ITEM_CODE3	20	1000	통계항목3코드
    # ITEM_NAME3	200	총자산증가율	통계항목명3
    # UNIT_NAME	200	십억원	단위
    # TIME	8	201101	시점
    # DATA_VALUE	23	48083.1	값

    statisticSearchURL = f'{URL}/StatisticSearch/{key}/json/kr/1/10/{stat_code}/{cycle}/{start_time}/{end_time}/'
    statisticSearch = requests.get(statisticSearchURL).json()
    print(statisticSearch)
    print('\n')
    """

